import argparse
import random
import sys
from typing import Generator, List, Tuple

import openpyxl
from openpyxl import load_workbook
import logging
import logging.handlers

logger = logging.getLogger(__name__)
handler = logging.handlers.RotatingFileHandler("create_class.log", maxBytes=10000, backupCount=5)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)

logger.addHandler(handler)
logger.addHandler(stream_handler)

verbose = False
quiet = False


def read_file(path: str) -> Generator:
    """
    Lies excel file zeilenweise aus (yield)
    :param path:
    :return:
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue
        yield [row[0].value, row[1].value, row[2].value]

def replace_umlaut(s: str) -> str:
    """
    ersetzt umlaute
    :param s:
    :return:
    """
    replacements = {
        'ä': 'ae',
        'ö': 'oe',
        'ü': 'ue',
        'Ä': 'Ae',
        'Ö': 'Oe',
        'Ü': 'Ue',
        'ß': 'ss'
    }
    for umlaut, replacement in replacements.items():
        s = s.replace(umlaut, replacement)
    return s

def escape(s: str) -> str:
    """
    ersetzt ' und "
    :param s:
    :return:
    """
    return s.replace('"', '\\"').replace("'", "\\'").replace("`", "\\`")

def create_user(i: List[str], file, pwd: str) -> None:
    """
    Erstellt Befehle die user erstellen
    :param i:
    :param file:
    :param pwd:
    :return:
    """
    logger.info(f"Creating user: {i[0]}")
    if verbose:
        print(f"echo creating user: {i[0]}", file=file)
    command1 = (
        f"getent passwd k{replace_umlaut(str(i[0]).lower())} > /dev/null && echo 'User " 
        f"{replace_umlaut(str(i[0]).lower())} already exists. Aborting.' && exit 1 || true"
    )
    command2 = f"groupadd {replace_umlaut(str(i[0]).lower())}"

    command3 = (
        f"useradd -d /home/klassen/{replace_umlaut(str(i[0]).lower())} "
        f"-c k{replace_umlaut(str(i[0]).lower())} "
        f"-m -g {replace_umlaut(str(i[0]).lower())} "
        f"-G cdrom,plugdev,sambashare -s /bin/bash k{replace_umlaut(str(i[0]).lower())}"
    )

    command4 = f"echo {replace_umlaut(str(i[0]).lower())}:{escape(pwd)} | chpasswd"

    print(command1, file=file)
    print(command2, file=file)
    print(command3, file=file)
    print(command4, file=file)

def delete_user(i: List[str], file) -> None:
    """
    Schreibt Befehle um user zu löschen
    :param i:
    :param file:
    :return:
    """
    logger.info(f"Deleting user: {i[0]}")
    if verbose: print(f"echo deleting user: {i[0]}", file=file)
    command = "userdel -r k" + replace_umlaut(str(i[0]).lower())
    print(command, file=file)

def generate_password(class_name: str, room: str, kv: str) -> str:
    """
    generiert pwd
    :param class_name:
    :param room:
    :param kv:
    :return:
    """
    logger.info("Generating password")
    chars = "1%&(),._-=^#"
    random_chars = [random.choice(chars) for _ in range(3)]
    password = f"{class_name}{random_chars[0]}{room}{random_chars[1]}{kv}{random_chars[2]}"
    return password



def create_credentials() -> Tuple[openpyxl.workbook.workbook.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """
    Excel wird erstellt um cred zu speichern
    :return:
    """
    logger.info("Creating credentials sheet")
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "Username"
    sheet["B1"] = "Password"
    return workbook, sheet

def add_credentials(sheet: openpyxl.worksheet.worksheet.Worksheet, name: str, row: int, pwd: str) -> None:
    """
    Credentials werden in Excel geschrieben
    :param sheet:
    :param i:
    :param row:
    :param pwd:
    :return:
    """
    logger.info(f"Adding credentials for user: {name}")
    sheet[f"A{row}"] = name
    sheet[f"B{row}"] = pwd

def save_credentials(workbook: openpyxl.workbook.workbook.Workbook) -> None:
    """
    speichert Excel
    :param workbook:
    :return:
    """
    logger.info("Saving credentials to file")
    workbook.save("user_credentials.xlsx")

def create_user_by_name(username: str, create_file, delete_file) -> None:
    """
    schreib create und del commands in file
    :param username:
    :param create_file:
    :param delete_file:
    :return:
    """
    logger.info(f"Creating user: {username}")
    if verbose: print(f"echo creating user: {username}", file=create_file)

    command1 = f"getent passwd {username} > /dev/null && echo 'User {username} already exists. Aborting.' && exit 1 || true"
    command2 = f"groupadd {username}"
    command3 = f"useradd -d /home/{username} -c {username} -m -g {username} -s /bin/bash {username}"
    command4 = f"echo {username}:{username} | chpasswd"

    print(command1, file=create_file)
    print(command2, file=create_file)
    print(command3, file=create_file)
    print(command4, file=create_file)

    logger.info(f"Deleting user: {username}")
    if verbose: print(f"echo deleting user: {username}", file=delete_file)
    command = "userdel -r " + username
    print(command, file=delete_file)

def create_files(path: str) -> None:
    """
    erstellt notwendige Files für user und iteriert durch alle user
    :param path:
    :return:
    """
    logger.info("Starting file creation")
    worksheet, sheet = create_credentials()
    row = 4
    with open("create_user.sh", "w") as create_user_file, open("delete_user.sh", "w") as delete_user_file:
        print("set -e", file=create_user_file)
        print("set -e", file=delete_user_file)
        print("mkdir /home/klassen", file=create_user_file)

        create_user_by_name("lehrer", create_user_file, delete_user_file)
        add_credentials(sheet, "lehrer", 2, "lehrer")
        create_user_by_name("seminar", create_user_file, delete_user_file)
        add_credentials(sheet, "seminar", 3, "seminar")

        for i in read_file(path):
            if i[0] == None:
                continue
            pwd = generate_password(str(i[0]).lower(), str(i[1]).lower(), str(i[2]).lower())
            create_user(i, create_user_file, pwd)
            delete_user(i, delete_user_file)
            add_credentials(sheet, i[0], row, pwd)
            row += 1
    save_credentials(worksheet)
    logger.info("Files created")













